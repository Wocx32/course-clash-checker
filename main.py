import pickle
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter

with open('course-data.pickle', 'rb') as file:
    data = pickle.load(file)

# for i in data:
#     if 'BIOL 100 B' in i['name']:
#         print(i)

input_courses = input('Enter courses: ').split(',')
input_courses = [x.strip().upper() for x in input_courses]

courses = []

# print(input_courses)

for i in data:
    if i['name'] in input_courses:
        if i['lab']:
            i['name'] = i['name'] + ' Lab'
            courses.append(i)
            continue
        courses.append(i)

# print(courses)

# print(courses[0])
# check if there is clash between timings of each course
def check_clash(course1, course2):
    course1_days = ''.join(course1['days'])
    course2_days = ''.join(course2['days'])

    if course1_days in course2_days or course2_days in course1_days:
        if course1['start_time'] <= course2['start_time'] <= course1['end_time'] or course1['start_time'] <= course2['end_time'] <= course1['end_time']:
            return True
    return False

detected = []

track_detected = False

for i in courses:
    for j in courses:
        if i['name'] != j['name'] and check_clash(i, j):
            for k in detected:
                if i['name'] in k and j['name'] in k:
                    track_detected = True
                    break
                track_detected = False
            
            if not track_detected:
                detected.append((i['name'], j['name']))
                print(i['name'], 'clashes with',j['name'])

def create_spreadsheet():
    workbook = Workbook()
    sheet = workbook.active

    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

    # set every column width to 40 px
    # set every row height to 20 px
    for i in range(1, 6):
        sheet.column_dimensions[get_column_letter(i)].width = 40
        sheet.row_dimensions[i].height = 23


    sheet['B1'] = 'hllo'
    # sheet['B1'] = 'world'

    workbook.save('test.xlsx')

# create_spreadsheet()